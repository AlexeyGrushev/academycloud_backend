import openpyxl


def read_data_from_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        data = [line.strip().split(' - ') for line in file]
    return data


def write_to_excel(data, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Слово'
    sheet['B1'] = 'Часть речи'

    for row_index, (word, part_of_speech) in enumerate(data, start=2):
        sheet.cell(row=row_index, column=1).value = word
        sheet.cell(row=row_index, column=2).value = part_of_speech

    workbook.save(excel_file)


# Имя текстового файла, из которого будут считываться данные
text_file = 'russian_spch.txt'

# Имя файла Excel, в который будут записываться данные
excel_file = 'words.xlsx'

# Чтение данных из текстового файла
data = read_data_from_file(text_file)

# Запись данных в Excel файл
write_to_excel(data, excel_file)
